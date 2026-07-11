import sqlite3
from rich.console import Console
from rich.table import Table
from datetime import datetime
from openpyxl import Workbook
import os

class Area_aluno():
    def __init__(self, id):
        self.id = id

    def verificar_dados(self):
        conexao = sqlite3.connect('comando/database/dados.db')
        cursor = conexao.cursor()

        comando_sql = """SELECT * FROM alunos WHERE id = ?"""
        cursor.execute(comando_sql, (self.id,))
        dados = cursor.fetchall()

        t = Table(title="dados")
        t.add_column('id', justify='center')
        t.add_column('nome', justify='center')
        t.add_column('idade', justify='center')
        t.add_column('peso', justify='center')
        t.add_column('altura', justify='center')
        t.add_column('id_treino', justify='center')

        for dado in dados:
            t.add_row(str(dado[0]), str(dado[1]), str(dado[2]), str(dado[3]), str(dado[4]), str(dado[5]))

        console = Console()
        console.print(t)

        conexao.close()

    def marcar_presenca(self):
        data = datetime.today().strftime('%d/%m/%Y')

        conexao = sqlite3.connect('comando/database/dados.db')
        cursor = conexao.cursor()

        comando_sql = """SELECT data FROM presenca WHERE id_aluno = ?"""
        cursor.execute(comando_sql, (self.id,))
        dado_presente = cursor.fetchall()

        for datas in dado_presente:
            if datas[0] == data:
                conexao.close()
                return 'presença já cadastrada antes'
        else:
            comando_sql = """INSERT INTO presenca (id_aluno, data,presente) VALUES (?,?,?)"""
            cursor.execute(comando_sql, (self.id, data, 'PRESENTE'))
            conexao.commit()
            conexao.close()
            return 'Cadastrado com sucesso'

    def ver_presencas(self):
        conexao = sqlite3.connect('comando/database/dados.db')
        cursor = conexao.cursor()
        comando_sql = """SELECT data,presente FROM presenca WHERE id_aluno = ?"""
        cursor.execute(comando_sql, (self.id,))
        dados = cursor.fetchall()

        t = Table(title="dados")
        t.add_column('data', justify='center')
        t.add_column('presença', justify='center')

        for dado in dados:
            t.add_row(str(dado[0]), str(dado[1]))

        c = Console()
        c.print(t)

        conexao.close()

    def ver_evolucao(self):
        conexao = sqlite3.connect('comando/database/dados.db')
        cursor =  conexao.cursor()

        comando_sql = """SELECT id,peso,altura,observacao,data FROM evolucao WHERE id_aluno = ?"""
        cursor.execute(comando_sql, (self.id,))
        dados = cursor.fetchall()

        t = Table(title="dados")
        t.add_column('id', justify='center')
        t.add_column('peso', justify='center')
        t.add_column('altura', justify='center')
        t.add_column('observacao', justify='center')

        for dado in dados:
            t.add_row(str(dado[0]), str(dado[1]), str(dado[2]), str(dado[3]))

        c = Console()
        c.print(t)

        conexao.close()

    def ver_treino_semana(self):
        dias_semana = ['segunda', 'terça','quarta','quinta','sexta','sabado','domingo']

        conexao = sqlite3.connect('comando/database/dados.db')
        cursor = conexao.cursor()

        comando_sql = """SELECT id_treino FROM alunos WHERE id = ?"""
        cursor.execute(comando_sql, (self.id,))
        treino_aluno = cursor.fetchall()[0][0]


        comando_sql = """SELECT
        exercicios.nome,
        exercicios_treinos.dia_semana,
        exercicios_treinos.series,
        exercicios_treinos.repeticoes
        FROM exercicios_treinos
        JOIN exercicios
        ON exercicios_treinos.id_exercicio = exercicios.id
        WHERE exercicios_treinos.id_treino = ?
        
        ORDER BY exercicios_treinos.dia_semana;"""

        cursor.execute(comando_sql, (treino_aluno,))
        dado = cursor.fetchall()

        t = Table(title="exercicios_treinos")
        t.add_column('exercicios_treinos', justify='center')
        t.add_column('exercicios_treinos', justify='center')
        t.add_column('exercicios_treinos', justify='center')
        t.add_column('exercicios_treinos', justify='center')

        for dados in dado:
            t.add_row(str(dados[0]),str(dias_semana[dados[1]-1]),str(dados[2]),str(dados[3]))

        c = Console()
        c.print(t)

        conexao.close()

    def exportar_excel(self):

        try:

            conexao = sqlite3.connect('comando/database/dados.db')
            cursor = conexao.cursor()

            comando_sql = """SELECT data,presente FROM presenca WHERE id_aluno = ?"""
            cursor.execute(comando_sql, (self.id,))
            dados_presenca = cursor.fetchall()

            comando_sql = """SELECT peso,altura,observacao,data FROM evolucao WHERE id_aluno = ?"""
            cursor.execute(comando_sql, (self.id,))

            dados_evolucao = cursor.fetchall()

            planilha = Workbook()
            aba = planilha.active

            aba.title = "RELATÓRIO"

            aba['B2'] = 'DATA'
            aba['C2'] = 'ESTATOS'
            aba['E2'] = 'DATA'
            aba['F2'] = 'PESO'
            aba['G2'] = 'ALTURA'
            aba['H2'] = 'OBSERVACAO'

            for posicao, dados in enumerate(dados_presenca, start=2):
                data, estatos = dados

                aba.cell(row=posicao + 1, column=2, value=data)
                aba.cell(row=posicao + 1, column=3, value=estatos)

            for posicao,dados in enumerate(dados_evolucao, start=2):
                peso,altura,observacao,data = dados
                aba.cell(row=posicao + 1, column=5, value=data)
                aba.cell(row=posicao + 1, column=6, value=peso)
                aba.cell(row=posicao + 1, column=7, value=altura)
                aba.cell(row=posicao + 1, column=8, value=observacao)

            planilha.save('comando/database/relatorio.xlsx')
            conexao.close()
            print('Salvando em:', os.getcwd())
            return 'Salvo com sucesso'

        except:

            return 'Erro ao acessar dados'
